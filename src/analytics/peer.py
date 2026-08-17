from __future__ import annotations
from pathlib import Path
import sqlite3,numpy as np,pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Font
import matplotlib.pyplot as plt
ROOT=Path(__file__).resolve().parents[2];DB=ROOT/'db/nifty100.db';OUT=ROOT/'output';RAD=ROOT/'reports/radar_charts';RAD.mkdir(parents=True,exist_ok=True)
RANK_METRICS={'roe':('return_on_equity_pct',True),'roce':('return_on_capital_employed_pct',True),'net_profit_margin':('net_profit_margin_pct',True),'de':('debt_to_equity',False),'fcf':('free_cash_flow_cr',True),'pat_cagr_5yr':('pat_cagr_5yr',True),'revenue_cagr_5yr':('revenue_cagr_5yr',True),'eps_cagr_5yr':('eps_cagr_5yr',True),'interest_coverage':('interest_coverage',True),'asset_turnover':('asset_turnover',True)}
RADAR_AXES=['ROE','ROCE','NPM','D/E','FCF score','PAT CAGR 5Y','Revenue CAGR 5Y','Composite Score']
def percentile_rank(values,higher=True):
 s=pd.to_numeric(values,errors='coerce');r=s.rank(pct=True,method='average')*100;return r if higher else 100-r
def load_peer_frame(con):
 from src.screener.engine import latest_financials,add_composite
 return add_composite(latest_financials(con))
def compute_peer_percentiles(d):
 rows=[]
 for group,g in d.dropna(subset=['peer_group_name']).groupby('peer_group_name'):
  for metric,(col,higher) in RANK_METRICS.items():
   ranks=percentile_rank(g[col].replace([np.inf,-np.inf],np.nan),higher)
   for idx,rank in ranks.items():rows.append({'company_id':g.loc[idx,'company_id'],'peer_group_name':group,'metric':metric,'value':g.loc[idx,col],'percentile_rank':rank,'year':g.loc[idx,'anchor_year']})
 return pd.DataFrame(rows)
def write_percentiles(con,p):
 con.execute('DROP TABLE IF EXISTS peer_percentiles');con.execute('''CREATE TABLE peer_percentiles(id INTEGER PRIMARY KEY AUTOINCREMENT,company_id TEXT NOT NULL,peer_group_name TEXT NOT NULL,metric TEXT NOT NULL,value REAL,percentile_rank REAL,year TEXT,UNIQUE(company_id,peer_group_name,metric,year),FOREIGN KEY(company_id) REFERENCES companies(id))''');p.where(pd.notna(p),None).to_sql('peer_percentiles',con,if_exists='append',index=False);con.commit()
def radar_values(row,g):
 pairs=[('return_on_equity_pct',True),('return_on_capital_employed_pct',True),('net_profit_margin_pct',True),('debt_to_equity',False),('free_cash_flow_cr',True),('pat_cagr_5yr',True),('revenue_cagr_5yr',True),('composite_quality_score',True)];vals=[];avg=[]
 for col,higher in pairs:
  s=g[col].replace([np.inf,-np.inf],np.nan);x=row[col]
  if pd.isna(x):vals.append(0);avg.append(0);continue
  p10,p90=s.quantile(.10),s.quantile(.90)
  if pd.isna(p10) or pd.isna(p90) or p90==p10:vals.append(50);avg.append(50);continue
  v=float(np.clip(x,p10,p90));a=float(np.clip(s.mean(),p10,p90));
  if not higher:v,a=100-v,100-a
  vals.append((v-p10)/(p90-p10)*100);avg.append((a-p10)/(p90-p10)*100)
 return vals,avg
def make_radar(row,g,path):
 vals,avg=radar_values(row,g);ang=np.linspace(0,2*np.pi,8,endpoint=False).tolist();ang+=ang[:1];v=vals+[vals[0]];a=avg+[avg[0]];fig=plt.figure(figsize=(7,7));ax=fig.add_subplot(111,polar=True);ax.plot(ang,v,linewidth=2);ax.fill(ang,v,alpha=.18);ax.plot(ang,a,linestyle='--',linewidth=1.5);ax.set_xticks(ang[:-1]);ax.set_xticklabels(RADAR_AXES,fontsize=8);ax.set_ylim(0,100);ax.set_title(f"{row.company_name} — {row.peer_group_name}");ax.legend(['Company','Peer average'],loc='upper right',bbox_to_anchor=(1.25,1.15),fontsize=8);fig.tight_layout();fig.savefig(path,dpi=160,bbox_inches='tight');plt.close(fig)
def make_standalone(row,avg,path):
 val=float(row.return_on_equity_pct) if pd.notna(row.return_on_equity_pct) else 0;ref=float(avg) if pd.notna(avg) else 0;fig,ax=plt.subplots(figsize=(7,4));ax.bar(['Company','Nifty 100 avg'],[val,ref]);ax.set_ylabel('ROE %');ax.set_title(f"{row.company_name} — Standalone ROE");fig.tight_layout();fig.savefig(path,dpi=160);plt.close(fig)
def export_peer_comparison(d,p):
 from src.screener.engine import DISPLAY_METRICS
 path=OUT/'peer_comparison.xlsx'
 with pd.ExcelWriter(path,engine='openpyxl') as writer:
  for group,g in d.dropna(subset=['peer_group_name']).groupby('peer_group_name'):
   base=g[['company_id','company_name']+[c for _,c in DISPLAY_METRICS]].copy()
   for metric in RANK_METRICS:
    q=p[(p.peer_group_name==group)&(p.metric==metric)][['company_id','percentile_rank']];base=base.merge(q,on='company_id',how='left');base.rename(columns={'percentile_rank':metric+'_percentile'},inplace=True)
   base.to_excel(writer,sheet_name=group[:31],index=False)
 wb=load_workbook(path);green=PatternFill('solid',fgColor='C6EFCE');yellow=PatternFill('solid',fgColor='FFF2CC');red=PatternFill('solid',fgColor='FFC7CE');amber=PatternFill('solid',fgColor='F4B183')
 for group,g in d.dropna(subset=['peer_group_name']).groupby('peer_group_name'):
  ws=wb[group[:31]];headers={c.value:c.column for c in ws[1]};ranks=[h for h in headers if h.endswith('_percentile')]
  for r in range(2,ws.max_row+1):
   cid=ws.cell(r,1).value;bench=bool(g.loc[g.company_id==cid,'is_benchmark'].iloc[0]);
   if bench:
    for cell in ws[r]:cell.fill=amber
   for h in ranks:
    v=ws.cell(r,headers[h]).value
    if v is not None:ws.cell(r,headers[h]).fill=green if v>=75 else red if v<=25 else yellow
  sr=ws.max_row+2;ws.cell(sr,1).value='Peer Median';ws.cell(sr,1).font=Font(bold=True)
  for h,ci in headers.items():
   if h in ['company_id','company_name'] or h.endswith('_percentile') or h not in g.columns:continue
   vals=pd.to_numeric(g[h],errors='coerce');
   if vals.notna().any():ws.cell(sr,ci).value=float(vals.median())
  ws.freeze_panes='A2';ws.auto_filter.ref=ws.dimensions
 wb.save(path);return path
def run():
 con=sqlite3.connect(DB);d=load_peer_frame(con);p=compute_peer_percentiles(d);write_percentiles(con,p)
 for _,row in d.iterrows():
  safe=str(row.company_id).replace('/','_')
  if pd.notna(row.peer_group_name):make_radar(row,d[d.peer_group_name==row.peer_group_name],RAD/f'{safe}_radar.png')
  else:make_standalone(row,d.return_on_equity_pct.mean(),RAD/f'{safe}_radar.png')
 path=export_peer_comparison(d,p);con.close();return p,path
if __name__=='__main__':p,path=run();print(f'peer_percentiles={len(p)} rows; workbook={path}')
