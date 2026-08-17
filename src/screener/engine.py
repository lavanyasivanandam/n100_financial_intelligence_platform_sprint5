from __future__ import annotations
from pathlib import Path
import argparse, sqlite3
import numpy as np
import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
ROOT=Path(__file__).resolve().parents[2]
DB_PATH=ROOT/'db/nifty100.db'; CONFIG=ROOT/'config/screener_config.yaml'; OUTPUT=ROOT/'output'; OUTPUT.mkdir(exist_ok=True)
FILTER_COLUMNS={'roe_min':'return_on_equity_pct','de_max':'debt_to_equity','fcf_min':'free_cash_flow_cr','revenue_cagr_5yr_min':'revenue_cagr_5yr','pat_cagr_5yr_min':'pat_cagr_5yr','opm_min':'operating_profit_margin_pct','pe_max':'pe_ratio','pb_max':'pb_ratio','dividend_yield_min':'dividend_yield_pct','icr_min':'interest_coverage','market_cap_min':'market_cap_crore','net_profit_min':'net_profit_cr','eps_cagr_5yr_min':'eps_cagr_5yr','asset_turnover_min':'asset_turnover','sales_min':'sales_cr','revenue_cagr_3yr_min':'revenue_cagr_3yr'}
DISPLAY_METRICS=[('ROE %','return_on_equity_pct'),('ROCE %','return_on_capital_employed_pct'),('NPM %','net_profit_margin_pct'),('OPM %','operating_profit_margin_pct'),('D/E','debt_to_equity'),('ICR','interest_coverage'),('FCF Cr','free_cash_flow_cr'),('FCF CAGR 5Y %','fcf_cagr_5yr'),('PAT CAGR 5Y %','pat_cagr_5yr'),('Revenue CAGR 5Y %','revenue_cagr_5yr'),('EPS CAGR 5Y %','eps_cagr_5yr'),('Asset Turnover','asset_turnover'),('Net Profit Cr','net_profit_cr'),('Sales Cr','sales_cr'),('Market Cap Cr','market_cap_crore'),('P/E','pe_ratio'),('P/B','pb_ratio'),('Dividend Yield %','dividend_yield_pct'),('Composite Score','composite_quality_score'),('Broad Sector','broad_sector')]
COMPOSITE_COMPONENTS={'roe':('return_on_equity_pct',True,.15),'roce':('return_on_capital_employed_pct',True,.10),'npm':('net_profit_margin_pct',True,.10),'fcf_cagr':('fcf_cagr_5yr',True,.15),'cfo_pat':('cfo_pat_ratio',True,.10),'fcf_positive':('fcf_positive_flag',True,.05),'revenue_cagr':('revenue_cagr_5yr',True,.10),'pat_cagr':('pat_cagr_5yr',True,.10),'de':('debt_to_equity',False,.10),'icr':('interest_coverage',True,.05)}

def latest_financials(con):
    pl=pd.read_sql_query('SELECT * FROM profitandloss',con); fr=pd.read_sql_query('SELECT * FROM financial_ratios',con); cf=pd.read_sql_query('SELECT * FROM cashflow',con); mc=pd.read_sql_query('SELECT * FROM market_cap',con); sec=pd.read_sql_query('SELECT company_id,broad_sector AS sector_broad FROM sectors',con); comp=pd.read_sql_query('SELECT id company_id,company_name FROM companies',con); peers=pd.read_sql_query('SELECT company_id,peer_group_name,is_benchmark FROM peer_groups',con)
    pl['yn']=pl.year.str[:4].astype(int); fr['yn']=fr.year.str[:4].astype(int); cf['yn']=cf.year.str[:4].astype(int); anchor=pl.sort_values(['company_id','yn']).groupby('company_id').tail(1); fr2=fr.sort_values(['company_id','yn']); rows=[]
    for _,a in anchor.iterrows():
        h=fr2[(fr2.company_id==a.company_id)&(fr2.year<=a.year)].tail(1)
        if h.empty: continue
        row=h.iloc[0].to_dict(); row.update(company_id=a.company_id,anchor_year=a.year,yn=int(a.yn),sales_cr=a.sales,net_profit_cr=a.net_profit); rows.append(row)
    d=pd.DataFrame(rows); mc['year']=pd.to_numeric(mc.year,errors='coerce'); latest_mc=mc.sort_values(['company_id','year']).groupby('company_id').tail(1); d=d.merge(latest_mc[['company_id','year','market_cap_crore','pe_ratio','pb_ratio','dividend_yield_pct']],on='company_id',how='left',suffixes=('','_mc')); d=d.merge(comp,on='company_id',how='left').merge(sec,on='company_id',how='left').merge(peers,on='company_id',how='left'); d['broad_sector']=d['broad_sector'].fillna(d['sector_broad']) if 'broad_sector' in d.columns else d['sector_broad']; d.drop(columns=['sector_broad'],inplace=True)
    cf2=cf.sort_values(['company_id','yn']); cfo=[]
    for _,a in anchor.iterrows():
        h=cf2[(cf2.company_id==a.company_id)&(cf2.year<=a.year)].tail(1); cfo.append((a.company_id,np.nan if h.empty else h.iloc[0].operating_activity))
    d=d.merge(pd.DataFrame(cfo,columns=['company_id','cfo_cr']),on='company_id',how='left'); d['cfo_pat_ratio']=d.cfo_cr/d.net_profit_cr.replace(0,np.nan); d['fcf_positive_flag']=(d.free_cash_flow_cr>0).astype(float)
    fr2['fcf_year']=fr2.year.str[:4].astype(int); fcf_map=fr2[['company_id','fcf_year','free_cash_flow_cr']].drop_duplicates(['company_id','fcf_year'])
    def fcagr(r):
        q=fcf_map[(fcf_map.company_id==r.company_id)&(fcf_map.fcf_year==r.yn-5)]
        if q.empty or pd.isna(q.iloc[0].free_cash_flow_cr) or pd.isna(r.free_cash_flow_cr): return np.nan
        s,e=float(q.iloc[0].free_cash_flow_cr),float(r.free_cash_flow_cr)
        if s<=0 or e<=0:return np.nan
        return ((e/s)**(1/5)-1)*100
    d['fcf_cagr_5yr']=d.apply(fcagr,axis=1); return d

def winsor_scale(s,higher=True):
    x=pd.to_numeric(s,errors='coerce').replace([np.inf,-np.inf],np.nan)
    if x.notna().sum()==0:return pd.Series(50.,index=s.index)
    p10,p90=x.quantile(.10),x.quantile(.90)
    if pd.isna(p10) or pd.isna(p90) or p90==p10:return pd.Series(50.,index=s.index)
    score=(x.clip(p10,p90)-p10)/(p90-p10)*100; return (score if higher else 100-score).fillna(50.)

def add_composite(d):
    d=d.copy()
    for key,(col,higher,_) in COMPOSITE_COMPONENTS.items():
        vals=pd.Series(index=d.index,dtype=float)
        for _,g in d.groupby('broad_sector',dropna=False):vals.loc[g.index]=winsor_scale(g[col],higher)
        d['_score_'+key]=vals
    d['composite_quality_score']=sum(d['_score_'+k]*w for k,(_,_,w) in COMPOSITE_COMPONENTS.items()).round(2); return d

def apply_filters(d,filters,db_path=DB_PATH):
    mask=pd.Series(True,index=d.index)
    for key,threshold in filters.items():
        if key=='de_declining':continue
        col=FILTER_COLUMNS.get(key)
        if not col:continue
        if key=='de_max':cond=(d[col]<=float(threshold))|(d.broad_sector.fillna('')=='Financials')
        elif key=='icr_min':cond=d[col].where(d.interest_coverage_label!='Debt Free',np.inf)>=float(threshold)
        elif key.endswith('_max'):cond=pd.to_numeric(d[col],errors='coerce')<=float(threshold)
        else:cond=pd.to_numeric(d[col],errors='coerce')>=float(threshold)
        mask &= cond.fillna(False)
    if filters.get('de_declining'):
        con=sqlite3.connect(db_path); full=pd.read_sql_query('SELECT company_id,year,debt_to_equity FROM financial_ratios',con); con.close(); full['yn']=full.year.str[:4].astype(int); full=full.sort_values(['company_id','yn'])
        def dec(g):g=g.tail(2);return len(g)==2 and pd.notna(g.iloc[0].debt_to_equity) and pd.notna(g.iloc[1].debt_to_equity) and g.iloc[1].debt_to_equity<g.iloc[0].debt_to_equity
        trend=full.groupby('company_id').apply(dec).to_dict();mask &= d.company_id.map(trend).fillna(False)
    return d.loc[mask].sort_values('composite_quality_score',ascending=False).copy()

def load_config(path=CONFIG):
    with open(path,'r',encoding='utf-8') as f:return yaml.safe_load(f)
def run_preset(name,d,config):return apply_filters(d,config['presets'][name])
def export_screener(results,config):
    path=OUTPUT/'screener_output.xlsx'
    with pd.ExcelWriter(path,engine='openpyxl') as writer:
        for name,df in results.items():df.reindex(columns=['company_id','company_name']+[c for _,c in DISPLAY_METRICS]).to_excel(writer,sheet_name=name[:31],index=False)
    wb=load_workbook(path);green=PatternFill('solid',fgColor='C6EFCE');red=PatternFill('solid',fgColor='FFC7CE')
    for name,df in results.items():
        ws=wb[name[:31]];headers={c.value:c.column for c in ws[1]};filters=config['presets'][name]
        for i in range(2,ws.max_row+1):
            rec=df.iloc[i-2]
            for key,v in filters.items():
                if key=='de_declining':continue
                col=FILTER_COLUMNS.get(key)
                if not col or col not in headers:continue
                val=rec.get(col);ok=False
                if key=='de_max' and rec.get('broad_sector')=='Financials':ok=True
                elif key=='icr_min' and rec.get('interest_coverage_label')=='Debt Free':ok=True
                elif pd.notna(val):ok=float(val)<=float(v) if key.endswith('_max') else float(val)>=float(v)
                ws.cell(i,headers[col]).fill=green if ok else red
        ws.freeze_panes='A2';ws.auto_filter.ref=ws.dimensions
    wb.save(path);return path
if __name__=='__main__':
    ap=argparse.ArgumentParser(); ap.add_argument('--preset'); ap.add_argument('--threshold',action='append',default=[],help='Override a threshold, e.g. --threshold roe_min=20'); args=ap.parse_args()
    cfg=load_config()
    con=sqlite3.connect(DB_PATH); d=add_composite(latest_financials(con)); con.close()
    if args.preset:
        filters=dict(cfg['presets'].get(args.preset,{}))
        if args.preset=='custom' and not filters: filters={}
        for item in args.threshold:
            k,v=item.split('=',1)
            filters[k]=True if v.lower()=='true' else False if v.lower()=='false' else float(v)
        result=apply_filters(d,filters); print(result[['company_id','company_name','composite_quality_score']].to_string(index=False))
    else:
        results={n:run_preset(n,d,cfg) for n in cfg['presets']}; export_screener(results,cfg)
